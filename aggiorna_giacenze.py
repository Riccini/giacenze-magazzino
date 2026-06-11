import re, json, sys
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
from collections import defaultdict

FOGLIO_MAP = {
    "Corrugar Rotoli":          ("Corrugar",   "Rotoli"),
    "Corrugar Rotoli Drenaggio":("Corrugar",   "Rotoli Drenaggio"),
    "Corrugar Drenaggio":       ("Corrugar",   "Barre Drenaggio"),
    "Corrugar Barre":           ("Corrugar",   "Barre"),
    "Sedici Plus":              ("SediciPlus", "Tubi"),
    "Kingcor":                  ("Kingcor",    "Tubi"),
    "Tripplo ":                 ("Tripplo",    "Tubi"),
    "Monopipe":                 ("Monopipe",   "Tubi"),
    "PE 100 Gas Barre":         ("Polier",     "Barre Gas"),
    "PE 100 Gas Rotoli":        ("Polier",     "Rotoli Gas"),
    "Fluid Superfluid":         ("Superfluid", "Tubi"),
    "PE 100 Barre":             ("Polier",     "Barre PE100"),
    "PE 100 Rotoli":            ("Polier",     "Rotoli PE100"),
    "RC2":                      ("Polier",     "Rotoli RC2"),
    "Monotubo ":                ("Polier",     "Monotubo"),
    "BD IIP":                   ("Polier",     "Rotoli BD IIP"),
    "BD Irriga":                ("Polier",     "Rotoli BD Irriga"),
}

def get_var(p):
    d = p["d"].upper(); c = p["c"].upper(); m = p["macro"]; t = p["tipo"]
    if m == "Corrugar" and t == "Rotoli":
        l = "25m" if any(x in d for x in ["25M","M 25","M25","ML 25"]) else "50m"
        if "GIALLO" in d or c.startswith("19C"): return "Giallo " + l
        if "BLU IMQ" in d: return ("Blu c/tir. " if any(x in d for x in ["C/T","T.LO"]) else "Blu IMQ ") + l
        if "GRIGIO" in d: return "Grigio " + l
        if "NERO"   in d: return "Nero " + l
        if "ROSSO"  in d: return "Rosso " + l
        if "VERDE"  in d: return ("Verde FW " if "FASTWEB" in d else "Verde ") + l
        return p["d"][:20]
    if m == "Corrugar" and t == "Barre":
        col = "Nero" if "NERO" in d else ("Rosso" if "ROSSO" in d else "Grigio")
        res = "750N" if "750N" in d else "450N"
        lu = "3m" if any(x in d for x in [" 3M","M.3"]) or d.endswith("3M") else "6m"
        return res + " " + col + " " + lu
    if m == "Corrugar" and t == "Rotoli Drenaggio":
        l2 = "25m" if any(x in d for x in ["25M","M25","M 25","ML 25"]) else "50m"
        return ("Drenofilter" if "DRENOFILTER" in d else "Drenocor") + " " + l2
    if m == "Corrugar" and t == "Barre Drenaggio":
        lu = "3m" if any(x in d for x in [" 3M","M.3"]) else ("5m" if " 5M" in d else "6m")
        return "Drenobar " + lu
    if m == "Polier" and t == "Barre Gas":
        s = "S5" if "S5" in d else ("S8" if "S8" in d else "")
        lu = "12m" if any(x in d for x in ["12M","12 M"]) else "6m"
        return s + (" RC" if "RC" in d else "") + " " + lu
    if m == "Polier" and t == "Rotoli Gas":
        s = "S5" if "S5" in d else ("S8" if "S8" in d else "")
        mm = re.search(r"(\d+)\s*M\b", d)
        return s + (" RC" if "RC" in d else "") + " " + (mm.group(0).replace(" ","").lower() if mm else "")
    if m == "Polier" and t == "Barre PE100":
        pn = re.search(r"PN\s*([\d.,]+)", d); ps = "PN" + pn.group(1) if pn else ""
        lu = "12m" if c.endswith("D") or any(x in d for x in ["12M","11,8","11.8"]) else "6m"
        if any(x in d for x in ["10,3","10.3"]): lu = "10,3m"
        rc = " RC2" if "RC2" in d else (" RC" if "RC" in d else "")
        return (ps + rc + " " + lu).strip()
    if m == "Polier" and t == "Rotoli PE100":
        pn = re.search(r"PN\s*([\d.,]+)", d); ps = "PN" + pn.group(1) if pn else ""
        mm = re.search(r"(\d+)\s*M\b", d)
        rc = " RC2" if "RC2" in d else (" RC" if "RC" in d else "")
        return (ps + rc + " " + (mm.group(0).replace(" ","").lower() if mm else "100m")).strip()
    if m == "Polier" and t == "Rotoli RC2":
        pn = re.search(r"PN\s*([\d.,]+)", d); ps = "PN" + pn.group(1) if pn else ""
        is_b = c.startswith("06B")
        if is_b:
            lu = "12m" if c.endswith("D") or "12M" in d else "6m"
            return ("B " + ps + " " + lu).strip()
        else:
            mm = re.search(r"\b(25|50|100|200)\s*[Mm]\b", d)
            lu = mm.group(0).replace(" ","").lower() if mm else ("50m" if float(p["diam"]) >= 90 else "100m")
            return ("R " + ps + " " + lu).strip()
    if m == "Polier" and t == "Rotoli Idropolier":
        pn = re.search(r"PN\s*(\d+)", d); mm = re.search(r"(\d+)\s*M\b", d)
        return (("PN" + pn.group(1) if pn else "") + " " + (mm.group(0).replace(" ","").lower() if mm else "")).strip()
    if m == "Polier" and t in ("Rotoli BD IIP","Rotoli BD Irriga"):
        pn = re.search(r"PN\s*(\d+[\.,]?\d*)", d)
        mm = re.search(r"M\.\s*(\d+)|(\d+)\s*M\b", d)
        lu = ((mm.group(1) or mm.group(2)) + "m") if mm else ""
        return (("PN" + pn.group(1) if pn else "") + " " + lu).strip()
    if m == "Polier" and t == "Monotubo":
        sdr = re.search(r"SDR\s*([\d.,]+)", d)
        col = "N" if "NERO" in d else ("R" if "REDLINE" in d or "ROSSO" in d else "")
        mm = re.search(r"(\d{2,4})\s*M\b", d)
        return " ".join(x for x in [("SDR"+sdr.group(1) if sdr else "")+("R" if "RIG" in c else ""), col, (mm.group(0).replace(" ","").lower() if mm else "")] if x)
    if m in ("Monopipe","Tripplo"):
        sn = re.search(r"SN\s*(\d+)", d)
        lu = "3m" if any(x in d for x in [" 3M","M.3"]) or d.endswith("3M") else "6m"
        return (("SN" + sn.group(1) if sn else "") + " " + lu).strip()
    if m == "Kingcor":
        sn = re.search(r"SN\s*(\d+)", d)
        lu = "3m" if any(x in d for x in [" 3M","M.3"]) else "6m"
        return (("SN" + sn.group(1) if sn else "") + " " + lu).strip() or p["d"][:15]
    if m == "Superfluid":
        sn = re.search(r"SN\s*(\d+)", d)
        lu = "3m" if "3M" in d or "M.3" in d else "6m"
        ex = (" F4S" if "F4S" in d else "") + (" F6N" if "F6N" in d else "") + (" TNT" if "TNT" in d else "")
        return (("SN" + sn.group(1) if sn else "") + ex + " " + lu).strip()
    if m == "SediciPlus":
        sn = re.search(r"SN\s*(\d+)", d)
        lu = "3m" if "3M" in d or "M.3" in d else "6m"
        ex = (" F4S" if "F4S" in d else "") + (" F6N" if "F6N" in d else "")
        return (("SN" + sn.group(1) if sn else "") + ex + " " + lu).strip()
    return p["d"][:15]

# ── LISTA ARTICOLI ────────────────────────────────────────────────
wb_lista = load_workbook("Lista Articoli per consultazione.xlsx", read_only=True)
codici_lista = {}
for sn in wb_lista.sheetnames:
    if sn not in FOGLIO_MAP: continue
    macro, tipo_foglio = FOGLIO_MAP[sn]
    for row in wb_lista[sn].iter_rows(min_row=2, values_only=True):
        c = str(row[0] or "").strip()
        diam = str(row[2] or "").strip() if row[2] is not None else ""
        desc = str(row[1] or "").strip().upper()
        if not c: continue
        if macro == "Corrugar" and tipo_foglio in ("Rotoli","Barre"):
            tipo = "Barre" if "BARRA CORRUGAR" in desc else "Rotoli"
        else:
            tipo = tipo_foglio
        if c in codici_lista:
            if codici_lista[c]["tipo"] == "Barre" and tipo == "Rotoli":
                codici_lista[c] = {"macro": macro, "tipo": tipo, "diam": diam}
        else:
            codici_lista[c] = {"macro": macro, "tipo": tipo, "diam": diam}
print(f"Lista articoli: {len(codici_lista)} codici")

# ── SALDI ─────────────────────────────────────────────────────────
saldi_file = next(Path(".").glob("Stampa_saldi*.xlsx"), None)
if not saldi_file: print("ERRORE: Stampa_saldi non trovato"); sys.exit(1)
print(f"Saldi: {saldi_file}")
wb_saldi = load_workbook(saldi_file, read_only=True)
data = []
for row in wb_saldi.active.iter_rows(min_row=2, values_only=True):
    c = str(row[0] or "").strip()
    if c not in codici_lista: continue
    try: i = int(float(row[8] or 0))
    except: i = 0
    try: l = int(float(row[11] or 0))
    except: l = 0
    info = codici_lista[c]
    data.append({"c":c,"d":str(row[7] or "").strip(),"um":"ML",
                 "i":i,"l":l,"macro":info["macro"],"tipo":info["tipo"],"diam":info["diam"]})
print(f"Articoli: {len(data)}")

# ── PORTAFOGLIO ORDINI ────────────────────────────────────────────
ord_file = next(Path(".").glob("Stampa_portafoglio*.xlsx"), None)
ordini_art = defaultdict(list)
clienti_dict = {}
ordini_cli = defaultdict(lambda: defaultdict(lambda: {"righe":[], "totale":0.0}))

if ord_file:
    print(f"Ordini: {ord_file}")
    for row in load_workbook(ord_file, read_only=True).active.iter_rows(min_row=2, values_only=True):
        cod_cli  = str(row[1]  or "").strip()
        nome_cli = str(row[7]  or "").strip()
        num_doc  = str(row[2]  or "").strip()
        data_ord = row[4]
        cod_art  = str(row[12] or "").strip()
        desc_art = str(row[13] or "").strip()
        um       = str(row[15] or "").strip()
        try: qta    = float(row[16] or 0)
        except: qta = 0
        try: prezzo = float(row[18] or 0)
        except: prezzo = 0
        try: tot_riga = float(row[6] or 0)
        except: tot_riga = 0
        if not cod_cli or not nome_cli or qta == 0: continue
        data_str = data_ord.strftime("%d/%m/%Y") if hasattr(data_ord,"strftime") else str(data_ord or "")[:10]
        clienti_dict[cod_cli] = nome_cli
        ordini_art[cod_art].append({"cli":nome_cli,"ord":num_doc,"dat":data_str,"qta":round(int(qta))})
        ordini_cli[cod_cli][num_doc]["righe"].append(
            {"art":cod_art,"des":desc_art,"dat":data_str,"qta":round(int(qta)),"um":um,"prez":round(prezzo,4)})
        ordini_cli[cod_cli][num_doc]["totale"] += tot_riga
    print(f"Clienti: {len(clienti_dict)}")

# ── VARIANTI + DEDUP ──────────────────────────────────────────────
for p in data:
    p["var"] = get_var(p)
    p["ord"] = ordini_art.get(p["c"], [])
groups = defaultdict(list)
for p in data: groups[(p["macro"],p["tipo"],p["diam"],p["var"])].append(p)
for key, prods in groups.items():
    if len(prods) > 1:
        for p in prods: p["var"] = p["var"] + " (" + p["c"][-4:] + ")"

# ── GENERA JS ─────────────────────────────────────────────────────
def esc(s): return s.replace("\\","\\\\").replace('"','\\"')

lines = []
for p in data:
    lines.append('{'
        + 'c:"'+esc(p["c"])+'"'
        + ',d:"'+esc(p["d"])+'"'
        + ',i:'+str(p["i"])+',l:'+str(p["l"])
        + ',"macro":"'+esc(p["macro"])+'"'
        + ',"tipo":"'+esc(p["tipo"])+'"'
        + ',"diam":"'+esc(p["diam"])+'"'
        + ',"var":"'+esc(p["var"])+'"'
        + ',"ord":'+json.dumps(p["ord"], ensure_ascii=False)
        + '}')
data_js = "var DATA=[\n" + ",\n".join(lines) + "\n];"

cli_list = sorted([{"cod":k,"nome":v} for k,v in clienti_dict.items()], key=lambda x: x["nome"])
cli_js = "var CLIENTI=[\n" + ",\n".join(
    ['{"cod":"'+esc(c["cod"])+'","nome":"'+esc(c["nome"])+'"}' for c in cli_list]
) + "\n];"

ord_cli_parts = []
for cod_cli, ordini in ordini_cli.items():
    ord_list = []
    for num_ord, dati in sorted(ordini.items()):
        dat = dati["righe"][0]["dat"] if dati["righe"] else ""
        ord_list.append({"ord":num_ord,"dat":dat,"tot":round(dati["totale"],2),"righe":dati["righe"]})
    ord_cli_parts.append('"'+esc(cod_cli)+'":'+json.dumps(ord_list, ensure_ascii=False))
ordini_js = "var ORDINI_CLI={\n" + ",\n".join(ord_cli_parts) + "\n};"

now = datetime.now().strftime("%d/%m/%Y %H:%M")

# ── AGGIORNA TEMPLATE ─────────────────────────────────────────────
with open("index_template.html", encoding="utf-8") as f:
    html = f.read()

html = re.sub(r"var DATA=\[.*?\];", data_js, html, flags=re.DOTALL)
html = re.sub(r"var CLIENTI=\[.*?\];", cli_js, html, flags=re.DOTALL)
html = re.sub(r"var ORDINI_CLI=\{.*?\};", ordini_js, html, flags=re.DOTALL)
html = re.sub(r'Ultimo aggiornamento: <strong[^>]*>[^<]*</strong>',
              'Ultimo aggiornamento: <strong id="update-date">'+now+'</strong>', html)
html = re.sub(r'<span class="topbar-badge">[^<]*</span>',
              f'<span class="topbar-badge">{len(data)} articoli</span>', html)

with open("index.html", "w", encoding="utf-8") as f:
    f.write(html)
print(f"OK: {len(data)} art, {len(cli_list)} clienti — {now}")

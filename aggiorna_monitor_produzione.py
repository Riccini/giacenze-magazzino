import json, sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

# ── LISTA ARTICOLI (anagrafica: codice -> categoria/diametro/nome) ──
wb_lista = load_workbook("Lista Articoli per consultazione.xlsx", read_only=True, data_only=True)
anagrafica = {}
duplicati_ignorati = []
for sheet in wb_lista.sheetnames:
    ws = wb_lista[sheet]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        codice = str(row[0]).strip()
        nome_vis = (row[3] or "").strip() if len(row) > 3 and row[3] else ""
        entry = {
            "categoria": sheet.strip(),
            "descrizione": row[1],
            "diametro": row[2],
            "nome_visualizzare": nome_vis
        }
        # Alcuni codici compaiono per errore in piu' fogli (es. un rotolo
        # duplicato anche nel foglio Barre, con la riga incompleta). In
        # quel caso teniamo sempre la voce con il nome da visualizzare
        # compilato, a prescindere dall'ordine dei fogli nel file.
        if codice not in anagrafica:
            anagrafica[codice] = entry
        elif not anagrafica[codice]["nome_visualizzare"] and nome_vis:
            anagrafica[codice] = entry
        elif anagrafica[codice]["categoria"] != sheet.strip():
            duplicati_ignorati.append((codice, anagrafica[codice]["categoria"], sheet.strip()))

print(f"Lista articoli: {len(anagrafica)} codici")
if duplicati_ignorati:
    print(f"Duplicati tra fogli ignorati ({len(duplicati_ignorati)}): {duplicati_ignorati}")

# ── SALDI ARTICOLI (giacenza SLQTAPER, impegnato SLQTIPER, ordinato SLQTOPER) ──
saldi_file = next(Path(".").glob("Stampa_saldi*.xlsx"), None)
if not saldi_file:
    print("ERRORE: file Stampa_saldi*.xlsx non trovato")
    sys.exit(1)
print(f"Saldi: {saldi_file}")

wb_saldi = load_workbook(saldi_file, read_only=True, data_only=True)
ws_saldi = wb_saldi.active
headers_saldi = [c.value for c in next(ws_saldi.iter_rows(min_row=1, max_row=1))]
idx_codice = headers_saldi.index("SLCODICE")
idx_giacenza = headers_saldi.index("SLQTAPER")
idx_impegnato = headers_saldi.index("SLQTIPER")
idx_ordinato = headers_saldi.index("SLQTOPER")

saldi = {}
for row in ws_saldi.iter_rows(min_row=2, values_only=True):
    codice = str(row[idx_codice] or "").strip()
    if not codice or codice in saldi:
        continue
    saldi[codice] = {
        "giacenza": row[idx_giacenza] or 0,
        "impegnato": row[idx_impegnato] or 0,
        "ordinato": row[idx_ordinato] or 0,
    }
print(f"Articoli in saldi: {len(saldi)}")

# ── PORTAFOGLIO ORDINI (ordini aperti per articolo, con data di evasione) ──
ord_file = next(Path(".").glob("Stampa_portafoglio*.xlsx"), None)
ordini_per_articolo = defaultdict(list)
if ord_file:
    print(f"Ordini: {ord_file}")
    wb_ord = load_workbook(ord_file, read_only=True, data_only=True)
    ws_ord = wb_ord.active
    headers_ord = [c.value for c in next(ws_ord.iter_rows(min_row=1, max_row=1))]
    idx_art = headers_ord.index("ARCODART")
    idx_cliente = headers_ord.index("ANDESCRI")
    idx_qta = headers_ord.index("MVQTAMOV")
    idx_evasione = headers_ord.index("MVDATEVA")

    for row in ws_ord.iter_rows(min_row=2, values_only=True):
        codice = str(row[idx_art] or "").strip()
        cliente = row[idx_cliente]
        if not codice or not cliente:
            continue
        data_ev = row[idx_evasione]
        data_str = data_ev.strftime("%Y-%m-%d") if hasattr(data_ev, "strftime") else (str(data_ev)[:10] if data_ev else None)
        ordini_per_articolo[codice].append({
            "cliente": cliente,
            "quantita": row[idx_qta] or 0,
            "data_evasione": data_str
        })
else:
    print("ATTENZIONE: file Stampa_portafoglio*.xlsx non trovato, procedo senza dettaglio ordini")


import re

def calc_pct(giacenza, impegnato, disponibile):
    if giacenza and giacenza != 0:
        return round((disponibile / giacenza) * 100, 1)
    elif impegnato and impegnato > 0:
        return -100.0
    return 0.0


def extract_rating(nome):
    """Estrae il valore numerico di classi tipo SN8, PN16, 450N, SDR11 per l'ordinamento orizzontale."""
    if not nome:
        return (999, "")
    m = re.match(r'^(\d+)\s*N\b', nome)
    if m:
        return (int(m.group(1)), nome)
    m = re.match(r'^[A-Za-z]+(\d+)', nome)
    if m:
        return (int(m.group(1)), nome)
    return (999, nome)


# ── COSTRUISCO STRUTTURA categoria -> diametro -> articoli ──
categorie_dict = defaultdict(lambda: defaultdict(list))
for codice, info in anagrafica.items():
    s = saldi.get(codice, {"giacenza": 0, "impegnato": 0, "ordinato": 0})
    giacenza = s["giacenza"]
    impegnato = s["impegnato"]
    disponibile = giacenza - impegnato

    ordini = sorted(
        ordini_per_articolo.get(codice, []),
        key=lambda o: (o["data_evasione"] is None, o["data_evasione"])
    )

    articolo_obj = {
        "codice": codice,
        "descrizione": info["descrizione"],
        "nome_visualizzare": info["nome_visualizzare"],
        "giacenza": giacenza,
        "impegnato": impegnato,
        "ordinato": s["ordinato"],
        "disponibile": disponibile,
        "pct": calc_pct(giacenza, impegnato, disponibile),
        "ordini": ordini
    }
    diametro_key = info["diametro"] if info["diametro"] is not None else "N/D"
    categorie_dict[info["categoria"]][diametro_key].append(articolo_obj)

categorie_output = []
for categoria, diametri_dict in categorie_dict.items():
    diametri_output = []
    cat_giacenza = cat_impegnato = cat_disponibile = 0
    cat_allerta = False
    cat_critici = 0

    for diametro, articoli in sorted(diametri_dict.items(), key=lambda x: (isinstance(x[0], str), x[0])):
        d_giacenza = sum(a["giacenza"] for a in articoli)
        d_impegnato = sum(a["impegnato"] for a in articoli)
        d_disponibile = d_giacenza - d_impegnato
        d_allerta = d_disponibile < 0

        diametri_output.append({
            "diametro": diametro,
            "giacenza": d_giacenza,
            "impegnato": d_impegnato,
            "disponibile": d_disponibile,
            "allerta": d_allerta,
            "pct": calc_pct(d_giacenza, d_impegnato, d_disponibile),
            "articoli": sorted(articoli, key=lambda a: extract_rating(a["nome_visualizzare"] or a["descrizione"] or ""))
        })
        cat_giacenza += d_giacenza
        cat_impegnato += d_impegnato
        cat_disponibile += d_disponibile
        if d_allerta:
            cat_allerta = True
        cat_critici += sum(1 for a in articoli if a["disponibile"] < 0)

    n_articoli = sum(len(a) for a in diametri_dict.values())
    categorie_output.append({
        "categoria": categoria,
        "giacenza": cat_giacenza,
        "impegnato": cat_impegnato,
        "disponibile": cat_disponibile,
        "allerta": cat_allerta,
        "pct": calc_pct(cat_giacenza, cat_impegnato, cat_disponibile),
        "n_articoli": n_articoli,
        "n_critici": cat_critici,
        "diametri": sorted(diametri_output, key=lambda d: (isinstance(d["diametro"], str), d["diametro"]))
    })

categorie_output.sort(key=lambda c: c["disponibile"])

report_critici = []
for categoria, diametri_dict in categorie_dict.items():
    for diametro, articoli in diametri_dict.items():
        for a in articoli:
            if a["disponibile"] < 0:
                report_critici.append({**a, "categoria": categoria, "diametro": diametro})
report_critici.sort(key=lambda a: a["disponibile"])

now = datetime.now().strftime("%d/%m/%Y %H:%M")
output = {"generato": now, "categorie": categorie_output, "report_critici": report_critici}

# ── INSERISCO NEL TEMPLATE ──
with open("monitor_produzione_template.html", encoding="utf-8") as f:
    html = f.read()

idx = html.index("const DATA = ")
idx_end = html.index(";", idx) + 1
data_js = "const DATA = " + json.dumps(output, ensure_ascii=False) + ";"
html = html[:idx] + data_js + html[idx_end:]

with open("monitor_produzione.html", "w", encoding="utf-8") as f:
    f.write(html)

print(f"OK: {len(anagrafica)} articoli, {len(report_critici)} critici — {now}")
